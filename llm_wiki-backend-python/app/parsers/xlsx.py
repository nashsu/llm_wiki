"""XLSX / XLS / ODS spreadsheet parser using openpyxl."""

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.parsers.base import DocumentParser, ParseResult
from app.parsers.registry import register_parser

logger = logging.getLogger(__name__)


def _worksheet_to_markdown(ws) -> str:
    """Convert an ``openpyxl`` worksheet into a Markdown table string.

    The sheet name is emitted as a ``##`` heading, followed by the data as a
    pipe-formatted table.  Empty rows are skipped; merged-cell content is
    taken from the top-left cell of the merge range.
    """
    lines: list[str] = []
    sheet_name = ws.title
    lines.append(f"## {sheet_name}")

    # Gather the cell range that actually contains data
    if ws.max_row is None or ws.max_column is None or ws.max_row < 1 or ws.max_column < 1:
        lines.append("*(empty sheet)*\n")
        return "\n".join(lines)

    # Resolve merged-cell values: openpyxl reads only the top-left cell
    # of a merged range as None for the other cells.
    merged_map: dict[tuple[int, int], object] = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds  # 1-based
        # Cache the value from the top-left cell
        top_left = ws.cell(row=min_row, column=min_col)
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                merged_map[(row_idx, col_idx)] = top_left.value

    def cell_value(row: int, col: int) -> str:
        """Get the display value of a cell, handling merged cells."""
        if (row, col) in merged_map:
            val = merged_map[(row, col)]
        else:
            val = ws.cell(row=row, column=col).value
        if val is None:
            return ""
        # Format numbers / dates as needed
        return str(val)

    # Collect data rows (skip completely empty rows)
    data_rows: list[list[str]] = []
    for row_idx in range(1, ws.max_row + 1):
        row_data = [cell_value(row_idx, col_idx) for col_idx in range(1, ws.max_column + 1)]
        if any(cell.strip() for cell in row_data):
            data_rows.append(row_data)

    if not data_rows:
        lines.append("*(empty sheet)*\n")
        return "\n".join(lines)

    col_count = max(len(r) for r in data_rows)

    # Normalise widths
    normalised = []
    for row in data_rows:
        while len(row) < col_count:
            row.append("")
        normalised.append(row)

    # Build Markdown table
    lines.append("| " + " | ".join(normalised[0]) + " |")
    lines.append("| " + " | ".join(["---"] * col_count) + " |")
    for row in normalised[1:]:
        lines.append("| " + " | ".join(row) + " |")

    lines.append("")  # trailing blank line
    return "\n".join(lines)


@register_parser()
class XLSXParser(DocumentParser):
    """Parser for Excel spreadsheets (``.xlsx``, ``.xls``, ``.ods``).

    Each worksheet is converted to a ``##``-headed Markdown table.
    """

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls", ".ods"]

    def parse(self, file_path: str) -> ParseResult:
        path = Path(file_path)
        try:
            wb = load_workbook(filename=str(path), data_only=True)
        except Exception as exc:
            logger.error("Failed to open spreadsheet %s: %s", path, exc)
            return ParseResult(
                text="",
                images=[],
                metadata={
                    "source": str(path),
                    "extension": path.suffix,
                    "parser": self.name,
                },
                success=False,
                error=f"Failed to open spreadsheet: {exc}",
            )

        try:
            text_parts: list[str] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                md = _worksheet_to_markdown(ws)
                text_parts.append(md)

            text = "\n".join(text_parts)

            metadata: dict = {
                "source": str(path),
                "extension": path.suffix,
                "parser": self.name,
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
            }

            return ParseResult(text=text, images=[], metadata=metadata)

        except Exception as exc:
            logger.error("Error parsing spreadsheet %s: %s", path, exc)
            return ParseResult(
                text="",
                images=[],
                metadata={
                    "source": str(path),
                    "extension": path.suffix,
                    "parser": self.name,
                },
                success=False,
                error=f"Error parsing spreadsheet: {exc}",
            )
        finally:
            wb.close()
